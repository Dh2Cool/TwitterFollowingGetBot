import tweepy
import pandas as pd
from openpyxl import load_workbook,Workbook
from os.path import exists
import os

consumer_key = "<api-consumer-key>" #from twitter api
consumer_secret = "<api-consumer-secret>"
access_token = "<access-token>"
access_token_secret = "<access-secret-token>"

auth = tweepy.OAuthHandler(consumer_key, consumer_secret)

auth.set_access_token(access_token, access_token_secret)

api = tweepy.API(auth)

screen_names = input("enter all names seperated by comma: \n")
screen_names1 = screen_names.split(',')

followers_id = []
followers = []
users_m = []
for x in range(len(screen_names1)):
    name = screen_names1[x]
    if not exists(f"{name}.xlsx"):
        wb = Workbook()
        wb.save(f"{name}.xlsx")
    else:
        load_workbook(f"{name}.xlsx")
    try:
        followers_id = api.get_friend_ids(screen_name=screen_names1[x],stringify_ids='true',count =100)
    except:
        print(f"{screen_names1[x]} cannot be found")
        os.remove(f"{screen_names1[x]}.xlsx")
        continue
    followers = api.lookup_users(user_id=followers_id)
    past_followers = pd.read_excel(f"{name}.xlsx")
    try:
        users_old = past_followers["Following:"].tolist()
    except:
        users_old = []
    users = []
    for user in followers:
        users.append(user.screen_name)
    user_24h = [ele for ele in users]
    for y in users:
        if y in users_old:
            user_24h.remove(y)
    for i in range(len(users) - len(user_24h)):
        user_24h.append(" ")
    col1 = "Following:"
    col2= "New following:"
    data = pd.DataFrame({col1:users,col2:user_24h})
    data.to_excel(f"{name}.xlsx")