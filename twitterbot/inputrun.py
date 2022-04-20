import tweepy
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook,Workbook
from os.path import exists

# assign the values accordingly
consumer_key = "otGx8bQYhUQi7md46ZTgoP4sl"
consumer_secret = "OAzkurTIwQhPFoGDdtVKxU8JvTg8Bw17iGp3NvSFyu4OeOjr99"
access_token = "767751198552186880-si284W5JFiBZsOpfgp1hGHFPlONK6Hy"
access_token_secret = "6YcnoPsuiW7blPR3pynU3nVFK9KSgX07HyBK7SZ9USQQ6"

# authorization of consumer key and consumer secret
auth = tweepy.OAuthHandler(consumer_key, consumer_secret)

# set access to user's access key and access secret
auth.set_access_token(access_token, access_token_secret)

# calling the api
api = tweepy.API(auth)





screen_names = input("enter all names seperated by comma")
screen_names1 = screen_names.split(',')

#wb.create_sheet("carnaugh1893")
followers_id = []
followers = []
users_m = []
print(screen_names1)
print(range(len(screen_names1)))
for x in range(len(screen_names1)):
    name = screen_names1[x]
    if not exists(f"{name}.xlsx"):
        wb = Workbook()
        wb.save(f"{name}.xlsx")
    else:
        load_workbook(f"{name}.xlsx")
    followers_id = api.get_friend_ids(screen_name=screen_names1[x],stringify_ids='true',count =100)
    followers = api.lookup_users(user_id=followers_id)
    past_followers = pd.read_excel(f"{name}.xlsx")
    try:
        users_old = past_followers["Followers:"].tolist()
    except:
        users_old = []
    users = []
    for user in followers:
        #print(user.screen_name)
        users.append(user.screen_name)
    user_24h = [ele for ele in users]
    print(users)
    print(users_old)
    for y in users:
        if y in users_old:
            user_24h.remove(y)
    for i in range(len(users) - len(user_24h)):
        user_24h.append(" ")
    print(user_24h)
    col1 = "Followers:"
    col2= "new followers"
    data = pd.DataFrame({col1:users,col2:user_24h})
    print(data)
    data.to_excel(f"{name}.xlsx")

    #users_m.append(users)
    #print("next loop")
#print(users_m)


#past_followers = pd.read_excel('tweepytrial.xlsx', sheet_name = 'sheet4')
